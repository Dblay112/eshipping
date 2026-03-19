"""Tally export views.

Mechanically extracted from views_old.py to keep behavior unchanged.
"""

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect

from openpyxl.styles import Font

from apps.tally.utils import _copy_row_style, _safe_date, _safe_str, build_tally_excel_from_template, workbook_to_bytes
from ..models import TallyInfo

from ._old_shared import _can_view_tally

import logging

logger = logging.getLogger(__name__)

@login_required(login_url='login')
def export_tally_excel(request, tally_id):
    """
    Export tally to Excel file using pre-formatted templates.

    Features:
    - Uses tally-type-specific Excel templates (bulk, straight_20, straight, japan)
    - Populates all tally header information (crop year, vessel, SD, tonnage, etc.)
    - Lists all containers with numbers, seals, bags, tonnage
    - Calculates totals (bags, tonnage)
    - For bulk tallies: includes expected/actual/saved bags with color coding
    - Dynamically inserts rows if more containers than template slots
    - Preserves template formatting and styling

    Security:
    - Permission check: only creator, supervisors, or managers can download
    - Uses _can_view_tally helper for authorization

    Permissions: Creator, terminal supervisors, managers, or superuser

    Args:
        tally_id: Primary key of TallyInfo to export

    Returns:
        Excel file download with populated tally data
        Or redirects to my_tallies if permission denied
    """
    tally = get_object_or_404(TallyInfo, id=tally_id)
    if not _can_view_tally(request.user, tally):
        messages.error(request, "You don't have permission to download this tally.")
        return redirect("my_tallies")

    template_map = {
        "BULK": "bulk.xlsx",
        "STRAIGHT_20FT": "straight_20.xlsx",
        "STRAIGHT_40FT": "straight.xlsx",
        "JAPAN_STRAIGHT_40FT": "japan.xlsx",
    }

    template_filename = template_map.get(tally.tally_type)
    if not template_filename:
        return HttpResponse(f"Unsupported tally_type: {tally.tally_type}", status=400)

    wb, ws, _ = build_tally_excel_from_template(
        template_filename=template_filename, sheet_name="NEW")

    bold = Font(bold=True)

    ws["C3"] = _safe_str(tally.crop_year)
    ws["G3"] = _safe_str(tally.produce or "RAW COCOA BEANS")

    ws["C4"] = _safe_str(tally.vessel)
    ws["G4"] = _safe_str(tally.sd_number)

    ws["C5"] = _safe_str(tally.terminal_name)
    ws["G5"] = float(tally.total_tonnage or 0)

    ws["C6"] = float(tally.total_tonnage or 0)

    if tally.tally_type == "STRAIGHT_20FT":
        ws["G6"] = "20FT"
    elif tally.tally_type in ["STRAIGHT_40FT", "JAPAN_STRAIGHT_40FT"]:
        ws["G6"] = "40FT"
    else:
        ws["G6"] = "BULK"

    ws["C7"] = _safe_str(tally.destination)
    ws["G7"] = _safe_date(tally.loading_date)

    ws["C8"] = _safe_str(tally.agent)
    ws["G8"] = _safe_str(tally.mk_number)

    ws["C9"] = _safe_str(tally.cocoa_type)
    ws["G9"] = _safe_str(tally.dry_bags)

    ws["C10"] = _safe_str(tally.superintendent_type)

    sup_names = tally.superintendent_name if isinstance(
        tally.superintendent_name, list) else []
    ws["C11"] = ", ".join([str(x) for x in sup_names if str(x).strip()])

    ws["G11"] = _safe_str(tally.loading_type or (
        "BULK" if tally.tally_type == "BULK" else "STRAIGHT"))

    start_row = 15
    base_clerk_row = 18

    containers = list(tally.containers.all())

    existing_slots = base_clerk_row - start_row
    needed = len(containers)
    extra_inserted = 0

    if needed > existing_slots:
        extra_inserted = needed - existing_slots
        insert_at = base_clerk_row
        ws.insert_rows(insert_at, amount=extra_inserted)
        for i in range(extra_inserted):
            _copy_row_style(ws, src_row=start_row,
                            dst_row=insert_at + i, max_col=10)

    clerk_row = base_clerk_row + extra_inserted

    total_bags = 0
    total_tonnage = 0.0

    for i, c in enumerate(containers):
        r = start_row + i

        if tally.tally_type == "BULK":
            bags_val = c.bags_cut if c.bags_cut is not None else (c.bags or 0)
            ton_val = float(c.tonnage or 0)
        else:
            bags_val = c.bags if c.bags is not None else 0
            if c.tonnage is not None:
                ton_val = float(c.tonnage)
            else:
                ton_val = float(bags_val) / 16.0

        ws[f"B{r}"] = _safe_str(c.container_number)
        ws[f"C{r}"] = int(bags_val)
        ws[f"D{r}"] = round(float(ton_val), 3)
        ws[f"E{r}"] = _safe_str(c.seal_number)

        total_bags += int(bags_val)
        total_tonnage += float(ton_val)

    ws[f"B{clerk_row}"] = "TOTAL"
    ws[f"C{clerk_row}"] = int(total_bags)
    ws[f"D{clerk_row}"] = round(float(total_tonnage), 3)

    ws[f"B{clerk_row}"].font = bold
    ws[f"C{clerk_row}"].font = bold
    ws[f"D{clerk_row}"].font = bold

    ws["G5"] = round(total_tonnage, 3)
    ws["C6"] = round(total_tonnage, 3)

    clerk_names = tally.clerk_name if isinstance(
        tally.clerk_name, list) else []
    ws[f"G{clerk_row}"] = ", ".join(
        [str(x) for x in clerk_names if str(x).strip()])

    if tally.tally_type == "BULK":
        r23 = 23 + extra_inserted
        r24 = 24 + extra_inserted
        r25 = 25 + extra_inserted

        exp = int(tally.expected_bags or 0)
        act = int(tally.actual_bags or 0)
        saved = int(tally.bags_saved or 0)

        ws[f"E{r23}"] = "EXPECTED BAGS:"
        ws[f"F{r23}"] = exp

        ws[f"E{r24}"] = "ACTUAL BAGS:"
        ws[f"F{r24}"] = act

        ws[f"E{r25}"] = "BAGS SAVED/LOST:"
        ws[f"F{r25}"] = saved

        ws[f"E{r23}"].font = bold
        ws[f"E{r24}"].font = bold
        ws[f"E{r25}"].font = bold

        ws[f"F{r23}"].font = bold
        ws[f"F{r24}"].font = bold

        if saved > 0:
            ws[f"F{r25}"].font = Font(bold=True, color="008000")
        elif saved < 0:
            ws[f"F{r25}"].font = Font(bold=True, color="FF0000")
        else:
            ws[f"F{r25}"].font = bold

    file_bytes = workbook_to_bytes(wb)

    filename = f"TALLY_{tally.tally_number}_{tally.tally_type}.xlsx"
    response = HttpResponse(
        file_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

