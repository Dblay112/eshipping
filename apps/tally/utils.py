import io
import os
import datetime
from copy import copy

import openpyxl
from django.conf import settings


def _safe_str(v):
    """
    Convert value to string safely, returning empty string for None.

    Used throughout tally export functions to handle None values gracefully
    when populating Excel and PDF templates.

    Args:
        v: Any value to convert to string

    Returns:
        str: Stripped string representation, or empty string if None

    Example:
        >>> _safe_str(None)
        ''
        >>> _safe_str('  hello  ')
        'hello'
        >>> _safe_str(123)
        '123'
    """
    if v is None:
        return ""
    return str(v).strip()


def _safe_date(v):
    """
    Convert value to Python date object for Excel export.

    Handles datetime objects by extracting date component. Returns empty string
    for None/empty values to prevent Excel errors.

    Args:
        v: Value to convert (datetime, date, or None)

    Returns:
        date: Python date object if valid
        str: Empty string if None/empty

    Example:
        >>> _safe_date(datetime.datetime(2026, 3, 19, 10, 30))
        datetime.date(2026, 3, 19)
        >>> _safe_date(datetime.date(2026, 3, 19))
        datetime.date(2026, 3, 19)
        >>> _safe_date(None)
        ''
    """
    if not v:
        return ""
    if isinstance(v, datetime.datetime):
        return v.date()
    return v  # date


def _copy_row_style(ws, src_row, dst_row, max_col=10):
    """
    Copy cell styles from source row to destination row in Excel worksheet.

    Preserves formatting when inserting new rows in tally Excel exports.
    Uses copy() to avoid shared style references across rows.

    Features:
    - Copies row height if set
    - Copies all cell styles (font, fill, border, alignment, protection)
    - Copies number format
    - Clears comments (prevents duplication)

    Args:
        ws: openpyxl worksheet object
        src_row: Source row number (1-indexed)
        dst_row: Destination row number (1-indexed)
        max_col: Maximum column to copy (default 10)

    Example:
        >>> _copy_row_style(ws, src_row=5, dst_row=6, max_col=10)
        # Copies styles from row 5 to row 6 for columns A-J
    """
    # row height (if set)
    if ws.row_dimensions.get(src_row) and ws.row_dimensions[src_row].height:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height

    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)

        if src.has_style:
            dst._style = copy(src._style)

        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        dst.comment = None


def build_tally_excel_from_template(
    *,
    template_filename: str,
    sheet_name: str = "NEW",
):
    """
    Load Excel template for tally export and return workbook objects.

    Loads pre-formatted Excel templates from apps/tally/excel_templates/
    directory. Templates contain styling and formulas for tally exports.

    Features:
    - Loads template from excel_templates directory
    - Selects specific sheet by name
    - Falls back to active sheet if sheet_name not found
    - Validates template exists before loading

    Args:
        template_filename: Filename of template (e.g., "bulk_tally_template.xlsx")
        sheet_name: Name of sheet to use (default "NEW")

    Returns:
        tuple: (workbook, worksheet, template_path)
            - workbook: openpyxl Workbook object
            - worksheet: openpyxl Worksheet object
            - template_path: Absolute path to template file

    Raises:
        FileNotFoundError: If template file doesn't exist

    Example:
        >>> wb, ws, path = build_tally_excel_from_template(
        ...     template_filename="bulk_tally_template.xlsx",
        ...     sheet_name="NEW"
        ... )
        >>> ws['A1'] = 'Tally Number'
    """
    template_path = os.path.join(
        settings.BASE_DIR, "apps", "tally", "excel_templates", template_filename
    )

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    return wb, ws, template_path


def workbook_to_bytes(wb) -> bytes:
    """
    Convert openpyxl workbook to bytes for HTTP response.

    Saves workbook to in-memory BytesIO buffer and returns raw bytes.
    Used for serving Excel files as downloads without writing to disk.

    Args:
        wb: openpyxl Workbook object

    Returns:
        bytes: Excel file content as bytes

    Example:
        >>> wb = openpyxl.Workbook()
        >>> excel_bytes = workbook_to_bytes(wb)
        >>> response = HttpResponse(
        ...     excel_bytes,
        ...     content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ... )
    """
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
